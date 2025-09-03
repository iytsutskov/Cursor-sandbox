from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

from src.application.use_cases.information_system_use_cases import ListInformationSystemsUseCase
from src.infrastructure.persistence.sqlite_information_system_repository import SQLiteInformationSystemRepository


class ExcelExportView(View):
    """View for exporting information systems to Excel"""
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.repository = SQLiteInformationSystemRepository()
    
    def get(self, request):
        """Export all information systems to Excel"""
        try:
            # Get all systems
            use_case = ListInformationSystemsUseCase(self.repository)
            response = use_case.execute()
            systems = response.systems
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Information Systems"
            
            # Define headers
            headers = [
                "ID", "Name", "Code", "Description", "Purpose", "Status", "System Type",
                "Owner Name", "Owner Email", "Owner Department", "Owner Phone",
                "Technology Stack", "Programming Languages", "Databases", "Frameworks",
                "Deployment Model", "Hosting Provider", "Business Functions",
                "Business Value", "Cost Center", "Version", "Parent System ID",
                "Dependent Systems", "Criticality Class", "Created At", "Updated At",
                "Dataflows Count", "Dataflows Details"
            ]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row, system in enumerate(systems, 2):
                # Basic system info
                ws.cell(row=row, column=1, value=str(system.id))
                ws.cell(row=row, column=2, value=system.name)
                ws.cell(row=row, column=3, value=system.code)
                ws.cell(row=row, column=4, value=system.description)
                ws.cell(row=row, column=5, value=system.purpose)
                ws.cell(row=row, column=6, value=system.status)
                ws.cell(row=row, column=7, value=system.system_type)
                
                # Owner info
                ws.cell(row=row, column=8, value=system.owner.name)
                ws.cell(row=row, column=9, value=system.owner.email)
                ws.cell(row=row, column=10, value=system.owner.department)
                ws.cell(row=row, column=11, value=system.owner.phone)
                
                # Technical spec
                tech_spec = system.technical_spec
                ws.cell(row=row, column=12, value=", ".join(tech_spec.technology_stack) if tech_spec.technology_stack else "")
                ws.cell(row=row, column=13, value=", ".join(tech_spec.programming_languages) if tech_spec.programming_languages else "")
                ws.cell(row=row, column=14, value=", ".join(tech_spec.databases) if tech_spec.databases else "")
                ws.cell(row=row, column=15, value=", ".join(tech_spec.frameworks) if tech_spec.frameworks else "")
                ws.cell(row=row, column=16, value=tech_spec.deployment_model)
                ws.cell(row=row, column=17, value=tech_spec.hosting_provider)
                
                # Business info
                business_functions = [f"{bf.name}: {bf.description}" for bf in system.business_functions]
                ws.cell(row=row, column=18, value="; ".join(business_functions) if business_functions else "")
                ws.cell(row=row, column=19, value=system.business_value)
                ws.cell(row=row, column=20, value=system.cost_center)
                ws.cell(row=row, column=21, value=system.version)
                ws.cell(row=row, column=22, value=str(system.parent_system_id) if system.parent_system_id else "")
                ws.cell(row=row, column=23, value=", ".join([str(sid) for sid in system.dependent_systems]) if system.dependent_systems else "")
                ws.cell(row=row, column=24, value=system.criticality_class)
                ws.cell(row=row, column=25, value=system.created_at)
                ws.cell(row=row, column=26, value=system.updated_at)
                
                # Dataflows info
                dataflows_count = len(system.dataflows) if system.dataflows else 0
                ws.cell(row=row, column=27, value=dataflows_count)
                
                # Dataflows details
                if system.dataflows:
                    dataflow_details = []
                    for df in system.dataflows:
                        detail = f"{df.source_system_id}→{df.target_system_id}: {df.data_objects} via {df.integration_technology}"
                        dataflow_details.append(detail)
                    ws.cell(row=row, column=28, value="; ".join(dataflow_details))
                else:
                    ws.cell(row=row, column=28, value="")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"information_systems_export_{timestamp}.xlsx"
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return HttpResponse(
                f'Failed to export to Excel: {str(e)}',
                status=500,
                content_type='text/plain'
            )
